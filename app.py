import streamlit as st
import json
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import io

# Konfigurasi halaman agar pas saat di-embed di Iframe
st.set_page_config(page_title="EduFinance Dashboard", layout="wide")

DAFTAR_BULAN = {
    "Januari": 31, "Februari": 28, "Maret": 31, "April": 30, 
    "Mei": 31, "Juni": 30, "Juli": 31, "Agustus": 31, 
    "September": 30, "Oktober": 31, "November": 30, "Desember": 31
}
URUTAN_BULAN = list(DAFTAR_BULAN.keys())

# ============================================================
# STATE MANAGEMENT (PENGGANTI DATA GLOBAL)
# ============================================================
if "data_keuangan" not in st.session_state:
    st.session_state.data_keuangan = {}
if "bulan_aktif" not in st.session_state:
    st.session_state.bulan_aktif = None

def simpan_data():
    with open("data.json", "w") as f:
        json.dump({
            "data_keuangan": st.session_state.data_keuangan,
            "bulan_aktif": st.session_state.bulan_aktif
        }, f, indent=2)

def load_data():
    if os.path.exists("data.json") and not st.session_state.data_keuangan:
        try:
            with open("data.json") as f:
                d = json.load(f)
                st.session_state.data_keuangan = d.get("data_keuangan", {})
                st.session_state.bulan_aktif = d.get("bulan_aktif", None)
        except:
            pass

load_data()

# Helper Hitung Saldo Akhir
def hitung_saldo_akhir(bulan):
    d = st.session_state.data_keuangan.get(bulan, {})
    saldo = d.get("saldo_awal", 0)
    masuk = sum(i["jumlah"] for h in d.get("pemasukan", {}).values() for i in h)
    keluar = sum(i["jumlah"] for h in d.get("pengeluaran", {}).values() for i in h)
    return saldo + masuk - keluar

def bulan_sebelumnya(bulan):
    try:
        idx = URUTAN_BULAN.index(bulan)
        if idx == 0: return None
        return URUTAN_BULAN[idx - 1]
    except:
        return None

# ============================================================
# TAMPILAN UTAMA & NAVIGASI MENU (8 MENU)
# ============================================================
st.title("💰 EduFinance Dashboard")

# Informasi Header Saldo Aktif
b_aktif = st.session_state.bulan_aktif
if b_aktif and b_aktif in st.session_state.data_keuangan:
    saldo_skrg = hitung_saldo_akhir(b_aktif)
    st.success(f"📌 **Bulan Kerja Aktif:** {b_aktif} | **Sisa Saldo Saat Ini:** Rp {saldo_skrg:,}")
else:
    st.warning("⚠️ Belum ada bulan aktif yang dipilih. Silakan atur di **Menu 1**.")

# Membuat Navigasi Menu menggunakan Tabs Streamlit
tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
    "1. Pilih Bulan", 
    "2. Input Saldo Awal", 
    "3. Tambah Transaksi", 
    "4. Ringkasan", 
    "5. Total Per Hari",
    "6. Edit / Hapus",
    "7. Export Excel",
    "8. Status Sistem"
])

# ============================================================
# MENU 1: PILIH BULAN
# ============================================================
with tab1:
    st.header("📅 Menu 1 — Pilih Bulan Kerja")
    bulan_terpilih = st.selectbox("Pilih nama bulan:", URUTAN_BULAN, key="sel_bulan")
    
    if st.button("Aktifkan & Buat Bulan", key="btn_aktifkan"):
        st.session_state.bulan_aktif = bulan_terpilih
        
        if bulan_terpilih not in st.session_state.data_keuangan:
            prev = bulan_sebelumnya(bulan_terpilih)
            saldo_carry = 0
            
            # Cek carry-over saldo otomatis dari bulan sebelumnya jika ada datanya
            if prev and prev in st.session_state.data_keuangan:
                saldo_carry = hitung_saldo_akhir(prev)
                st.session_state.carry_over_temp = saldo_carry
                st.info(f"Ada data bulan {prev} dengan saldo akhir Rp {saldo_carry:,}.")
            
            st.session_state.data_keuangan[bulan_terpilih] = {
                "saldo_awal": saldo_carry,
                "pemasukan": {},
                "pengeluaran": {}
            }
            simpan_data()
            st.success(f"Berhasil membuat record baru untuk bulan {bulan_terpilih}!")
        else:
            st.success(f"Bulan {bulan_terpilih} berhasil dipilih!")
        st.rerun()

# ============================================================
# MENU 2: INPUT SALDO AWAL
# ============================================================
with tab2:
    st.header("💵 Menu 2 — Input Saldo Awal")
    if not b_aktif:
        st.error("Pilih bulan dulu di Menu 1!")
    else:
        saldo_lama = st.session_state.data_keuangan[b_aktif].get("saldo_awal", 0)
        st.write(f"Saldo awal saat ini untuk bulan **{b_aktif}**: Rp {saldo_lama:,}")
        
        saldo_baru = st.number_input("Masukkan Saldo Awal Baru (Rp):", min_value=0, value=int(saldo_lama), step=50000)
        if st.button("Perbarui Saldo Awal"):
            st.session_state.data_keuangan[b_aktif]["saldo_awal"] = saldo_baru
            simpan_data()
            st.success(f"Saldo awal {b_aktif} berhasil diubah menjadi Rp {saldo_baru:,}!")
            st.rerun()

# ============================================================
# MENU 3: TAMBAH TRANSAKSI
# ============================================================
with tab3:
    st.header("📝 Menu 3 — Tambah Transaksi")
    if not b_aktif:
        st.error("Silakan tentukan bulan aktif terlebih dahulu!")
    else:
        max_hari = DAFTAR_BULAN[b_aktif]
        
        c1, c2 = st.columns(2)
        with c1:
            hari = st.number_input(f"Hari ke berapa? (1-{max_hari}):", min_value=1, max_value=max_hari, value=1)
            jenis = st.selectbox("Jenis Transaksi:", ["Pemasukan", "Pengeluaran"])
        with c2:
            keterangan = st.text_input("Keterangan / Deskripsi:")
            jumlah = st.number_input("Jumlah Uang (Rp):", min_value=0, value=0, step=10000)
            
        kategori = "kebutuhan"
        if jenis == "Pengeluaran":
            kategori = st.selectbox("Kategori Pengeluaran:", ["kebutuhan", "impulsif", "tabungan"])
            
        if st.button("Simpan Transaksi"):
            if not keterangan.strip():
                st.error("Keterangan tidak boleh kosong!")
            elif jumlah <= 0:
                st.error("Jumlah uang harus lebih besar dari Rp 0!")
            else:
                key = f"Day {hari}"
                if jenis == "Pemasukan":
                    st.session_state.data_keuangan[b_aktif]["pemasukan"].setdefault(key, []).append({
                        "keterangan": keterangan, "jumlah": jumlah
                    })
                    st.success(f"Sukses menambah Pemasukan: Rp {jumlah:,}")
                else:
                    st.session_state.data_keuangan[b_aktif]["pengeluaran"].setdefault(key, []).append({
                        "keterangan": keterangan, "kategori": kategori, "jumlah": jumlah
                    })
                    st.success(f"Sukses menambah Pengeluaran [{kategori}]: Rp {jumlah:,}")
                simpan_data()
                st.rerun()

# ============================================================
# MENU 4: RINGKASAN BULAN INI
# ============================================================
with tab4:
    st.header("📊 Menu 4 — Ringkasan Keuangan")
    if not b_aktif:
        st.error("Data ringkasan kosong, pilih bulan dahulu!")
    else:
        d = st.session_state.data_keuangan[b_aktif]
        saldo_aw = d["saldo_awal"]
        masuk = sum(i["jumlah"] for h in d["pemasukan"].values() for i in h)
        keluar = sum(i["jumlah"] for h in d["pengeluaran"].values() for i in h)
        akhir = saldo_aw + masuk - keluar
        
        kat_total = {"kebutuhan": 0, "impulsif": 0, "tabungan": 0}
        for h in d["pengeluaran"].values():
            for item in h:
                kat_total[item.get("kategori", "kebutuhan")] += item["jumlah"]
                
        # Tampilan Grid Metrik Finansial
        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        col_m1.metric("Saldo Awal", f"Rp {saldo_aw:,}")
        col_m2.metric("Total Masuk (+)", f"Rp {masuk:,}")
        col_m3.metric("Total Keluar (-)", f"Rp {keluar:,}")
        col_m4.metric("Saldo Akhir", f"Rp {akhir:,}")
        
        # Breakdown Kategori Pengeluaran
        st.subheader("🛒 Rincian Pengeluaran")
        cc1, cc2, cc3 = st.columns(3)
        cc1.info(f"🥦 **Kebutuhan:** Rp {kat_total['kebutuhan']:,}")
        cc2.warning(f"⚡ **Impulsif:** Rp {kat_total['impulsif']:,}")
        cc3.success(f"🐷 **Tabungan:** Rp {kat_total['tabungan']:,}")
        
        # Indikator Kesehatan Keuangan (Logika Persentase Status)
        if saldo_aw > 0:
            pct = (akhir / saldo_aw) * 100
            total_all_keluar = sum(kat_total.values())
            impct = (kat_total["impulsif"] / total_all_keluar * 100) if total_all_keluar > 0 else 0
            
            if pct >= 70: status, warna = "🟢 Aman", "green"
            elif pct >= 40: status, warna = "🟡 Waspada", "orange"
            elif pct >= 10: status, warna = "🟠 Hampir Habis", "orange"
            elif pct >= 0: status, warna = "🔴 Kritis", "red"
            else: status, warna = "❌ Defisit", "red"
            
            st.markdown(f"Sisa Saldo: **{pct:.1f}%** dari modal awal. Status Keuangan Anda: :{warna}[**{status}**]")
            if impct >= 50:
                st.error(f"⚠️ Bahaya! **{impct:.0f}%** pengeluaranmu bersifat impulsif! Rem belanjamu.")
            elif impct >= 30:
                st.warning(f"⚠️ **{impct:.0f}%** dompetmu bocor untuk hal impulsif. Mulai dikurangi ya.")

# ============================================================
# MENU 5: TOTAL PER HARI
# ============================================================
with tab5:
    st.header("📆 Menu 5 — Total per Hari")
    if not b_aktif:
        st.error("Pilih bulan terlebih dahulu.")
    else:
        d = st.session_state.data_keuangan[b_aktif]
        list_harian = []
        
        for i in range(1, DAFTAR_BULAN[b_aktif] + 1):
            key = f"Day {i}"
            m = sum(x["jumlah"] for x in d["pemasukan"].get(key, []))
            k = sum(x["jumlah"] for x in d["pengeluaran"].get(key, []))
            if m > 0 or k > 0:
                list_harian.append({"Hari": key, "Pemasukan (Rp)": m, "Pengeluaran (Rp)": k})
                
        if list_harian:
            st.table(pd.DataFrame(list_harian))
        else:
            st.info("Belum ada data transaksi yang dimasukkan di bulan ini.")

# ============================================================
# MENU 6: EDIT / HAPUS TRANSAKSI
# ============================================================
with tab6:
    st.header("⚙️ Menu 6 — Edit atau Hapus Transaksi")
    if not st.session_state.data_keuangan:
        st.info("Belum ada riwayat data keuangan terkumpul.")
    else:
        pilih_b = st.selectbox("Pilih Bulan Data:", list(st.session_state.data_keuangan.keys()), key="edit_b")
        pilih_j = st.radio("Pilih Jenis Data:", ["Pemasukan", "Pengeluaran"], horizontal=True)
        
        key_kat = pilih_j.lower()
        d_bulan = st.session_state.data_keuangan[pilih_b]
        
        # Kumpulkan list item transaksi
        semua_item = []
        for h, daftar_item in d_bulan[key_kat].items():
            for idx, item in enumerate(daftar_item):
                semua_item.append({"hari": h, "index_asli": idx, "keterangan": item["keterangan"], "jumlah": item["jumlah"], "item_ref": item})
                
        if not semua_item:
            st.write("Tidak ada data transaksi pada kategori ini.")
        else:
            df_view = pd.DataFrame([{"Hari": x["hari"], "Keterangan": x["keterangan"], "Jumlah": x["jumlah"]} for x in semua_item])
            st.dataframe(df_view, use_container_width=True)
            
            pilih_no = st.number_input("Pilih indeks nomor data transaksi (Mulai dari 0):", min_value=0, max_value=len(semua_item)-1, value=0, step=1)
            target = semua_item[pilih_no]
            
            c_e1, c_e2 = st.columns(2)
            with c_e1:
                edit_nominal = st.number_input("Tulis Jumlah Baru jika ingin edit (Rp):", min_value=1, value=int(target["jumlah"]))
                if st.button("Simpan Perubahan (Edit)"):
                    target["item_ref"]["jumlah"] = edit_nominal
                    simpan_data()
                    st.success("Berhasil mengubah nominal!")
                    st.rerun()
            with c_e2:
                st.write("Atau hapus baris transaksi ini permanen:")
                if st.button("🚨 Hapus Transaksi Ini"):
                    d_bulan[key_kat][target["hari"]].pop(target["index_asli"])
                    simpan_data()
                    st.success("Transaksi dihapus!")
                    st.rerun()

# ============================================================
# MENU 7: EXPORT EXCEL (MENGGUNAKAN LOGIKA OPENPYXL ASLI)
# ============================================================
with tab7:
    st.header("📥 Menu 7 — Export ke File Excel (.xlsx)")
    if not st.session_state.data_keuangan:
        st.error("Tidak ada data untuk diexport.")
    else:
        st.write("Klik tombol di bawah untuk membuat berkas Excel berdesain profesional berdasarkan data kamu.")
        
        # Logika pembentukan file Excel menggunakan Buffer Memory (agar bisa didownload langsung via browser)
        output = io.BytesIO()
        wb = Workbook()
        first = True
        
        def hf(c): return PatternFill("solid", fgColor=c)
        def tb():
            s = Side(style="thin", color="CCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)
        def sh(cell, bg="2C3E50", fg="FFFFFF"):
            cell.fill = hf(bg); cell.font = Font(bold=True, color=fg, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = tb()
        def sc(cell, align="left", num_fmt=None, bg=None):
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal=align, vertical="center"); cell.border = tb()
            if num_fmt: cell.number_format = num_fmt
            if bg: cell.fill = hf(bg)

        for bulan, data in st.session_state.data_keuangan.items():
            ws = wb.active if first else wb.create_sheet()
            ws.title = bulan[:15]
            first = False
            ws.sheet_view.showGridLines = False

            ws["A1"] = f"LAPORAN KEUANGAN — {bulan.upper()}"
            ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
            ws["A1"].fill = hf("1A5276")
            ws["A1"].alignment = Alignment(horizontal="center")
            ws.merge_cells("A1:D1")

            ws["A3"] = "PEMASUKAN"
            ws["A3"].font = Font(bold=True, color="FFFFFF")
            ws["A3"].fill = hf("27AE60")
            ws.merge_cells("A3:D3")

            for ci, h in enumerate(["Hari", "Keterangan", "Jumlah (Rp)"], 1):
                sh(ws.cell(4, ci, h), bg="1E8449")
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 35
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 18

            ri = 5
            total_masuk = 0
            for hari, items in data.get("pemasukan", {}).items():
                for item in items:
                    sc(ws.cell(ri, 1, hari), align="center")
                    sc(ws.cell(ri, 2, item["keterangan"]))
                    sc(ws.cell(ri, 3, item["jumlah"]), align="right", num_fmt="#,##0", bg="E8F8F5")
                    total_masuk += item["jumlah"]
                    ri += 1
            sc(ws.cell(ri, 2, "TOTAL PEMASUKAN"), align="right")
            sc(ws.cell(ri, 3, total_masuk), align="right", num_fmt="#,##0", bg="D5F5E3")
            ws.cell(ri, 2).font = Font(bold=True)
            ri += 2

            ws.cell(ri, 1, "PENGELUARAN").font = Font(bold=True, color="FFFFFF")
            ws.cell(ri, 1).fill = hf("E74C3C")
            ws.merge_cells(f"A{ri}:D{ri}")
            ri += 1

            for ci, h in enumerate(["Hari", "Keterangan", "Kategori", "Jumlah (Rp)"], 1):
                sh(ws.cell(ri, ci, h), bg="922B21")
            ri += 1

            total_keluar = 0
            kat_total = {"kebutuhan": 0, "impulsif": 0, "tabungan": 0}
            kc = {"kebutuhan": "D5F5E3", "impulsif": "FADBD8", "tabungan": "D6EAF8"}
            for hari, items in data.get("pengeluaran", {}).items():
                for item in items:
                    kat = item.get("kategori", "kebutuhan")
                    sc(ws.cell(ri, 1, hari), align="center")
                    sc(ws.cell(ri, 2, item["keterangan"]))
                    sc(ws.cell(ri, 3, kat.capitalize()), align="center", bg=kc.get(kat, "FFFFFF"))
                    sc(ws.cell(ri, 4, item["jumlah"]), align="right", num_fmt="#,##0", bg="FADBD8")
                    total_keluar += item["jumlah"]
                    kat_total[kat] += item["jumlah"]
                    ri += 1

            sc(ws.cell(ri, 3, "TOTAL PENGELUARAN"), align="right")
            sc(ws.cell(ri, 4, total_keluar), align="right", num_fmt="#,##0", bg="FADBD8")
            ws.cell(ri, 3).font = Font(bold=True)
            ri += 2

            saldo_awal = data.get("saldo_awal") or 0
            saldo_akhir = saldo_awal + total_masuk - total_keluar

            ws.cell(ri, 1, "RINGKASAN").font = Font(bold=True, color="FFFFFF")
            ws.cell(ri, 1).fill = hf("8E44AD")
            ws.merge_cells(f"A{ri}:D{ri}")
            ri += 1

            for label, val in [
                ("Saldo Awal", saldo_awal), ("Total Pemasukan", total_masuk),
                ("Total Pengeluaran", total_keluar), ("Saldo Akhir", saldo_akhir),
                ("  — Kebutuhan", kat_total["kebutuhan"]), ("  — Impulsif", kat_total["impulsif"]),
                ("  — Tabungan", kat_total["tabungan"])
            ]:
                sc(ws.cell(ri, 1, label))
                sc(ws.cell(ri, 2, val), align="right", num_fmt="#,##0", bg="D5F5E3" if val >= 0 else "FADBD8")
                if label == "Saldo Akhir":
                    ws.cell(ri, 1).font = Font(bold=True)
                    ws.cell(ri, 2).font = Font(bold=True)
                ri += 1

        ws2 = wb.create_sheet("Ringkasan Semua Bulan")
        ws2.sheet_view.showGridLines = False
        hdrs3 = ["Bulan", "Saldo Awal", "Pemasukan", "Pengeluaran", "Saldo Akhir", "Status"]
        wdths3 = [14, 18, 18, 18, 18, 14]
        for col, (h, w) in enumerate(zip(hdrs3, wdths3), 1):
            sh(ws2.cell(1, col, h), bg="1A5276")
            ws2.column_dimensions[get_column_letter(col)].width = w

        summary_rows = []
        for ri2, (bulan, data) in enumerate(st.session_state.data_keuangan.items(), 2):
            sa = data.get("saldo_awal") or 0
            m = sum(i["jumlah"] for h in data.get("pemasukan", {}).values() for i in h)
            k = sum(i["jumlah"] for h in data.get("pengeluaran", {}).values() for i in h)
            ak = sa + m - k
            pct = (ak / sa * 100) if sa > 0 else 0
            st_str = "Aman" if pct >= 70 else "Waspada" if pct >= 40 else "Hampir Habis" if pct >= 10 else "Defisit"
            bg_st = "D5F5E3" if st_str == "Aman" else "FDEBD0" if st_str == "Waspada" else "FADBD8"
            for ci, val in enumerate([bulan, sa, m, k, ak, st_str], 1):
                sc(ws2.cell(ri2, ci, val), align="center", num_fmt="#,##0" if ci in [2,3,4,5] else None, bg=bg_st if ci == 6 else ("F2F3F4" if ri2 % 2 == 0 else "FFFFFF"))
            summary_rows.append((bulan, m, k))

        wb.save(output)
        st.download_button(
            label="📊 Unduh Laporan Excel (.xlsx)",
            data=output.getvalue(),
            file_name="laporan_keuangan_edufinance.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ============================================================
# MENU 8: SIMPAN & STATUS SISTEM
# ============================================================
with tab8:
    st.header("💾 Menu 8 — Status Penyimpanan")
    st.info("Sistem web Streamlit menyimpan perubahan data kamu secara otomatis ke dalam berkas internal setiap kali transaksi ditambahkan atau diubah.")
    if st.button("Paksa Simpan Data Ke File (data.json)"):
        simpan_data()
        st.success("Data berhasil diamankan ke data.json!")
    st.write("Terima kasih telah menggunakan **EduFinance Dashboard!** 💰")
